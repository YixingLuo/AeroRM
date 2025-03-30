import re
import os
from openpyxl import Workbook, load_workbook
from extract_docx import process_document, extract_table
from extract_function import pipeline, search_functions

def find_files(path, key, filetype):
    file_list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            if key in file and file.endswith(filetype) and not file.startswith('~$'):
                file_list.append([root, file])
    return file_list

def find_key(file_dic, file):
    if file is None:
        return None
    for key in file_dic:
        if file == key:
            return key
    
    for key in file_dic:
        if file in key:
            return key
    
    return None

def get_content(file_dic, file, chap_number, chap_title):
    key = find_key(file_dic, file)
    if key is None:
        return None
    for item in file_dic[key]:
        if item['number'] == chap_number:
            return item['content']
        if chap_title and chap_title in item['title']:
            return item['content']
    return None

def write_to_excel(table, excel_path):
    wb = Workbook()
    ws = wb.active
    for i, row in enumerate(table, start=1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(excel_path)

def findPattern(text, patterns):
    if not text:
        return []
    results = []
    for pattern in patterns:
        tmp_results = re.findall(pattern, req_content1)
        for tmp_result in tmp_results:
            flag = True
            for result in results:
                if str(tmp_result) == str(result):
                    flag = False
                    break
            if flag:
                results.append(tmp_result)    
    return results

def addRef(file_dic, text, req_file):

    patterns = [
        r'(?:见|调用)(用户需求|任务书)?的?(\d(?:\.\d)*)', 
        r'《(.*?)》的?第?(\d(?:\.\d)*)', 
    ]
    if not text:
        return text
    ref_infos = findPattern(text, patterns)
    if not ref_infos:
        return text
    for ref_source, chap_number in ref_infos:
        if not ref_source:
            add_text = f'\n{chap_number}内容为:{get_content(file_dic, req_file, str(chap_number), None)}'
        else:
            add_text = f'\n{ref_source}的{chap_number}内容为:{get_content(file_dic, ref_source, str(chap_number), None)}'
        text += add_text
    return text

def findTestTable(table_list, req_id):
    # 基于正向跟踪矩阵的映射提取测试记录表
    # 1. 找到正向跟踪矩阵
    trace_table = None
    for table in table_list:
        if '需求规格说明中的需求项编号' in table and '对应的测试用例标识' in table:
            trace_table = table
            break
    if not trace_table or req_id not in trace_table: # 如果无跟踪矩阵或跟踪矩阵无req_id，就基于req_id找table
        for table in table_list:
            if req_id and req_id in table and '项目对应功能项' in table:
                return table
        return None
    # 2. 根据req_id找到测试用例标识
    test_ids = None
    for row in trace_table.split('\n'):
        if req_id in row:
            test_ids = row.split('|')[3]
            break
    # 3. 提取具有测试用例标识的table
    for table in table_list:
        for test_id in test_ids.split('\n'):
            if test_id and test_id in table and '项目对应功能项' in table:
                return table
    for table in table_list: # 如果还找不到，就基于req_id找table
        if req_id and req_id in table and '项目对应功能项' in table:
            return table
    return None

if __name__=='__main__':

    src_path = '设计组/空间站天和一号GNCC应用软件'
    IP_excel = '设计组/空间站天和一号GNCC应用软件/空间站天和一号GNCC应用软件-v1.xlsx'
    code_path = '设计组/空间站天和一号GNCC应用软件/4.代码工程/kjzth_gncc'

    wb = load_workbook(IP_excel)
    sheet = wb['IP列表']

    # 初始化
    ## 提取所有docx文章的所有章节，所有表格
    file_dic = {}
    docx_list = find_files(src_path, '.docx', '.docx')
    for root, file in docx_list:
        file_dic[file] = process_document(os.path.join(root, file), hastitle=True, recursion=True)
    ## 提取所有代码（鹤然）
    pipeline(code_path)

    ## 获取测试的table_list
    table_list = []
    for root, test_file in find_files(src_path, '确认测试', '.docx'):
        table_list += extract_table(os.path.join(root, test_file))

    # 比例
    total_num = 0
    req1_num = 0
    req2_num = 0
    design_num = 0
    code_num = 0
    test_num = 0

    can_generate_list = []

    # 遍历excel
    results_list = [['IP序号', '软件需求名称', '需求来源1', '需求来源2', '详细设计', '代码', '确认测试']]
    r = 4
    for row in sheet.iter_rows(min_row=4, values_only=True):
        total_num += 1
        if row[0] is None:
            break
        # 需求
        req_src1, req_chap1, req_src2, req_chap2 = row[4], row[5], row[7], row[6]
        req_content1 = get_content(file_dic, req_src1, str(req_chap1), None)
        req_content2 = get_content(file_dic, req_src2, str(req_chap2), None)
        if not req_content1:
            req1_num += 1
        if not req_content2:
            req2_num += 1
        # 引用问题
        req_content1 = addRef(file_dic, req_content1, req_src1)
        req_content2 = addRef(file_dic, req_content2, req_src2)


        # 设计
        design_files = find_files(src_path, '详细设计', '.docx')
        if design_files:
            _, design_file = design_files[0]
            func = row[3]
            # design_content = [func+'设计说明']
            # for item in file_dic[design_file]:
            #     if func+'设计说明' in item['title']:
            #         design_content += [item['title'], item['content']]
            # design_content = '\n'.join(design_content)
            if func:
                design_content = get_content(file_dic, design_file, None, func+'设计说明')
            else:
                design_content = ''
                design_num += 1
        else:
            design_content = ''
            design_num += 1


        # 代码
        functions = search_functions(row[3].replace('\n', ','))
        if isinstance(functions, list):
            functions = '\n'.join(functions)
        if not functions:
            code_num += 1

        # 测试
        key = row[2]
        test_result = findTestTable(table_list, key)
        if not test_result:
            test_num += 1
        
        results_list.append([row[0], row[1], req_content1, req_content2, design_content, functions, test_result])
        if (req_content1 or req_content2) and functions:
            can_generate_list.append('是')
        else:
            can_generate_list.append('否')
        sheet.cell(row=r, column=9).value = can_generate_list[-1]
        r += 1
    

    print(f'未提取出来：需求1 {req1_num}/{total_num}, 需求2 {req2_num}/{total_num}, 设计 {design_num}/{total_num}, 代码 {code_num}/{total_num}, 测试 {test_num}/{total_num}')
    write_to_excel(results_list, os.path.join(src_path, './result.xlsx'))
    sheet.cell(3,9).value='能否自动提取（是/否）'
    wb.save(IP_excel[:-6]+'-modified-2.xlsx')


