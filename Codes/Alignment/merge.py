import os
from openpyxl import Workbook, load_workbook

def write_to_excel(table, excel_path):
    wb = Workbook()
    ws = wb.active
    for i, row in enumerate(table, start=1):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    wb.save(excel_path)

files = [
   'file_path'     
]

infos = [['IP Name', 'Requirements', 'Implement']]
total_num = 0
complete_num = 0
for file in files:
    wb = load_workbook(file)
    sheet = wb['Sheet']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        IP_name = row[1]
        total_num += 1
        Req = '\n'.join([row[2] if row[2] else '', row[3] if row[3] else ''])
        Implement = row[5]
        if IP_name and Req.strip() and Implement:
            complete_num += 1
            infos.append([IP_name, Req, Implement])
write_to_excel(infos, './IP_Reqire.xlsx')
print(f'全部提取比例：{complete_num}/{total_num}')


infos = [['IP序号', '软件需求名称', '需求来源1', '需求来源2', '详细设计', '代码', '测试']]
total_num = 0
req1_num = 0
req2_num = 0
design_num = 0
code_num = 0
test_num = 0
complete_num = 0
for file in files:
    wb = load_workbook(file)
    sheet = wb['Sheet']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        infos.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6]])
        total_num += 1
        if not row[2]:
            req1_num += 1
        if not row[3]:
            req2_num += 1
        if not row[4]:
            design_num += 1
        if not row[5]:
            code_num += 1
        if not row[6]:
            test_num += 1
        if row[2] and row[3] and row[4] and row[5] and row[6]:
            complete_num += 1
print(f'提取比例：需求1 {total_num-req1_num}/{total_num}, 需求2 {total_num-req2_num}/{total_num}, 设计 {total_num-design_num}/{total_num}, 代码 {total_num-code_num}/{total_num}, 测试 {total_num-test_num}/{total_num}')
print(f'全部提取比例：{complete_num}/{total_num}')

write_to_excel(infos, './result.xlsx')

    
            